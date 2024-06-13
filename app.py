import smtplib
from email import encoders 
from email.mime.base import MIMEBase 
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import openpyxl


database = openpyxl.load_workbook(f"workbook.xlsx", data_only=True) #name of the file with the email IDs
database_sheet = database["Sheet1"] #put the name of the sheet with the emails

# uncomment out next line if an attachments
# filename = "filename.pdf" #attachment

fromAddress = "fromAddress@gmail.com" #from address
toAddress = "receiverAddress@gmail.com"
passw = "yourapppassword"  #app passkey to the from address
#set smtp server and port accoridng to your email provider
smtpServer = "smtp.gmail.com"
port = 587
notdone = []

#Login 
server = smtplib.SMTP(smtpServer, port)
server.ehlo()
server.starttls()
server.login(fromAddress, passw)
print("Login successful")
count = 0


# Sending mails one by one to emails in the sheet
for row in range(0,database_sheet.max_row + 1):
    try:
        if(database_sheet.cell(row, 1).value==None or database_sheet.cell(row, 1).value.strip()=="" or database_sheet.cell(row, 1).value.strip()=="None" ):
            break
        toAddress = database_sheet.cell(row, 1).value.strip()

        # create MIME object to add the body and attachment to
        msg = MIMEMultipart()
        msg["From"] = fromAddress
        msg["To"] = toAddress
        # msg["Bcc"] = "" #opitional



        msg["Subject"] = "Subject of the mail" #subject of the mail
        # attach the body of the mail as plain txt
        #body of the mail
        body = f"""
Hey there,

this a a sample mail

Thanks and regards
"""
        msg.attach(MIMEText(body, 'plain'))



        #uncomment out line 62 to 76 if there is an attachment
        # with open(filename, "rb") as attachment:
        #     # add file as an application/octet-stream
        #     # email client can then download this as an attachment
        #     part = MIMEBase("application", "octet-stream")
        #     part.set_payload(attachment.read())

        # # encode file
        # encoders.encode_base64(part)
        # part.add_header(
        #     "Content-Disposition",
        #     f"attachment; filename= {filename}"
        # )
        # #add message to the body and then convert it to a string
        # msg.attach(part)


 
        txt = msg.as_string()
        server.sendmail(fromAddress, toAddress, txt) #sending the mail
        print(f"Mail sent to {toAddress}")
        count += 1
    except:
        if(database_sheet.cell(row, 1).value==None or database_sheet.cell(row, 1).value.strip()=="" or database_sheet.cell(row, 1).value.strip()=="None" ):
            continue
        notdone.append(database_sheet.cell(row, 1).value) #adding email ID of all the undone mails in list

if(count != 1):
    print(f"{count} mails sent")
else:
    print("1 mail sent")
print()
print(notdone) #print out names of all the email IDs that emcountered errors
server.quit()
